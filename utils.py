# coding=utf-8
import xml.etree.ElementTree as ET
import openpyxl
from configparser import ConfigParser as cf
from openpyxl.styles import PatternFill
import difflib


class Utils(object):
    """
    获取xml文档的root节点
    """

    @staticmethod
    def get_xml_root(xml_path):
        return ET.parse(xml_path).getroot()

    """
     查找特定属性值标签的值
    """

    @staticmethod
    def get_attrib_value(root, attr_name):
        # keys = []
        for stu in root:
            # print(stu.attrib['name'])
            if stu.attrib['name'] == attr_name:
                # keys.append(attr_name)
                return stu.text
        return None

    """xml通过value来获取特性的name属性值"""

    """
    判断某个属性是否存在xml
    """

    @staticmethod
    def is_attr_name_exist(root, attr_name):
        for stu in root:
            if stu.attrib['name'] == attr_name:
                return True
        return False

    @staticmethod
    def get_key_by_value(root, value):
        keys = []
        for stu in root:
            if stu.text == value:
                keys.append(stu.attrib['name'])
        return keys

    """
        获取excel文件第一个sheet表格(已存在的excel)
    """

    @staticmethod
    def get_excel_sheet(wb, index):
        try:
            sheet = wb.worksheets[index]
            return sheet
        except Exception as e:
            print(e, "请检查wb/index参数")

        """
        获取已存在excel文件wb对象
        """

    @staticmethod
    def get_wb_exist(excel_path):
        try:
            return openpyxl.load_workbook(excel_path)
        except Exception as e:
            print(e, "请检查excel路径是否正确")

    """
    获取sheet页所最大行
    """

    @staticmethod
    def get_sheet_row(sheet):
        if sheet is not None:
            return sheet.max_row
        return 0

    """
       获取sheet页所最大列
    """

    @staticmethod
    def get_sheet_col(sheet):
        if sheet is not None:
            return sheet.max_column
        return 0

    """
          获取sheet页某个单元格的值
    """

    @staticmethod
    def get_cell_value(sheet, row, col):
        if sheet is not None:
            return sheet.cell(row, col).value
        return None

    """通过列字符和行获取值"""

    @staticmethod
    def get_cell_by_col_name(sheet, col_name, i):
        if sheet is not None:
            return sheet[col_name+str(i)].value
        return None
    """
    新建excel文件,wb_object对象
    """

    @staticmethod
    def create_excel_sheet():
        try:
            wb = openpyxl.Workbook()
            return wb
        except Exception as e:
            print(e, "创建wb_object失败")

    """
       保存wb
    """

    @staticmethod
    def save_wb(wb, file_name):
        wb.save(file_name)

    """
    写入数据
    """

    @staticmethod
    def write_sheet(sheet, data):
        if isinstance(data, list):
            sheet.append(data)
        else:
            print("传入参数必须是列表")

    """
      写入指定单元格数据
    """

    @staticmethod
    def write_sheet_cell(sheet, row, col, val):
        sheet.cell(row, col).value = val

    """填充单元格颜色"""

    @staticmethod
    def fill_color(sheet, i, col, result):
        if result == "FAIL":
            fill_color = "FF0000"
        elif result == "PASS":
            fill_color = "00FF00"
        else:
            fill_color = "0000FF"
        fill = PatternFill("solid", fgColor=fill_color)
        cell = sheet.cell(i, col)
        # print(cell)
        cell.fill = fill

    """
    获取配置文件对象cfg
    """

    @staticmethod
    def get_conf_object(conf_path, encoding="utf-8-sig"):
        try:
            cfg = cf()
            cfg.read(conf_path, encoding=encoding)
            return cfg
        except Exception as e:
            print(e)
            return None

    """
      获取配置文件对section下特定的key-value
    """

    @staticmethod
    def get_section_option(cfg, section, option):
        try:
            return cfg.get(section, option)
        except Exception as e:
            print(e)
            return None

    @staticmethod
    def get_difference(result, text1, text2):
        if result == "PASS":
            return "No"
        if result == "FAIL":
            d = difflib.Differ()
            text1 = text1.splitlines(keepends=True)
            text2 = text2.splitlines(keepends=True)
            difference = d.compare(text1, text2)
            return "\n".join(list(difference))
        else:
            return "BLOCK"

    @staticmethod
    def get_strings_flies_values(strings_path):
        try:
            values = list()
            with open(strings_path, encoding="utf-8") as f:
                line = f.readline()
                while line:
                    if "=" in line:
                        values.append(line)
                        # print(line)
                    line = f.readline()
            return values

        except Exception as e:
            print(e)
            return None

    @staticmethod
    def get_strings_key_by_vaule(strings_list, input_value):
        try:
            keys = list()
            for string_value in strings_list:
                key_value = string_value.split("=")
                key = eval(key_value[0].strip(""))
                value = eval(key_value[1].split(";")[0])
                if value == input_value:
                    keys.append(key)
            return keys
        except Exception as e:
            print(e)
            return list()
              
    @staticmethod
    def get_strings_value_by_key(strings_list, input_key):
        values = list()
        try:
            for string_value in strings_list:
                key_value = string_value.split("=")
                key = eval(key_value[0].strip(""))
                value = key_value[1].split(";")[0]
                if key == input_key:
                    values.append(eval(value))
            return values
        except Exception as e:
            print(e)
            return list()

